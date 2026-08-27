from fastapi import FastAPI, APIRouter, Request, Response, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import logging
import uuid
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone, timedelta
import httpx
import pandas as pd
import openpyxl

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI(title="PGP Course Change Request Portal")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ------------------------------------------------------------------
# Constants
# ------------------------------------------------------------------
ADMIN_EMAILS = {"secy.academics@iiml.ac.in"}
ALLOWED_DOMAINS = {"iim.ac.in", "iiml.ac.in"}
TERM_V_MIN_CREDITS = 5.0
TERM_V_MAX_CREDITS = 6.0
EMERGENT_SESSION_URL = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"

REQUEST_TYPES = {"ADD", "DROP", "ADD_DROP", "COURSE_SWAP", "SECTION_SWAP"}
TERMINAL_STATUSES = {"REJECTED", "PARTNER_REJECTED", "CANCELLED", "EXECUTED"}

def now_utc():
    return datetime.now(timezone.utc)

def iso(dt: datetime) -> str:
    return dt.isoformat()

# ------------------------------------------------------------------
# Auth helpers
# ------------------------------------------------------------------
async def get_session_token(request: Request) -> Optional[str]:
    token = request.cookies.get("session_token")
    if token:
        return token
    auth = request.headers.get("Authorization")
    if auth and auth.startswith("Bearer "):
        return auth.split(" ", 1)[1]
    return None

async def get_current_user(request: Request) -> Dict[str, Any]:
    token = await get_session_token(request)
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    expires_at = session["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < now_utc():
        raise HTTPException(status_code=401, detail="Session expired")
    user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user or not user.get("active", True):
        raise HTTPException(status_code=401, detail="User not found or inactive")
    return user

async def require_admin(request: Request) -> Dict[str, Any]:
    user = await get_current_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

async def require_student(request: Request) -> Dict[str, Any]:
    user = await get_current_user(request)
    if user.get("role") != "student":
        raise HTTPException(status_code=403, detail="Student access required")
    return user

# ------------------------------------------------------------------
# Audit
# ------------------------------------------------------------------
async def audit(action: str, actor: str, detail: str, request_id: Optional[str] = None):
    await db.audit_logs.insert_one({
        "log_id": f"log_{uuid.uuid4().hex[:12]}",
        "action": action,
        "actor": actor,
        "detail": detail,
        "request_id": request_id,
        "at": iso(now_utc()),
    })

# ------------------------------------------------------------------
# Auth endpoints
# ------------------------------------------------------------------
class SessionInput(BaseModel):
    session_id: str

@api_router.post("/auth/session")
async def create_session(payload: SessionInput, response: Response):
    async with httpx.AsyncClient(timeout=30) as hc:
        r = await hc.get(EMERGENT_SESSION_URL, headers={"X-Session-ID": payload.session_id})
    if r.status_code != 200:
        raise HTTPException(status_code=401, detail="Failed to validate session")
    data = r.json()
    email = data["email"].lower()
    domain = email.split("@")[-1]
    if domain not in ALLOWED_DOMAINS:
        raise HTTPException(status_code=403, detail=f"Only institutional accounts ({', '.join(ALLOWED_DOMAINS)}) are permitted.")

    # Determine role and pgpid
    is_admin = email in ADMIN_EMAILS
    role = "admin" if is_admin else "student"
    pgpid = None
    if not is_admin:
        student = await db.students.find_one({"email": email}, {"_id": 0})
        if student:
            pgpid = student["pgpid"]
        else:
            pgpid = email.split("@")[0].upper()

    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        user_id = existing["user_id"]
        await db.users.update_one({"user_id": user_id}, {"$set": {
            "name": data.get("name", existing.get("name")),
            "picture": data.get("picture"),
            "role": role,
            "pgpid": pgpid,
            "active": True,
        }})
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": user_id,
            "email": email,
            "name": data.get("name", email),
            "picture": data.get("picture"),
            "role": role,
            "pgpid": pgpid,
            "active": True,
            "created_at": iso(now_utc()),
        })

    session_token = data["session_token"]
    expires_at = now_utc() + timedelta(days=7)
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": iso(expires_at),
        "created_at": iso(now_utc()),
    })

    response.set_cookie(key="session_token", value=session_token, httponly=True,
                        secure=True, samesite="none", path="/", max_age=7*24*60*60)
    await audit("LOGIN", pgpid or email, f"{role} logged in")
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    user["session_token"] = session_token  # returned once so frontend can use Authorization fallback
    return user

@api_router.get("/auth/me")
async def auth_me(request: Request):
    return await get_current_user(request)

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    token = await get_session_token(request)
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    response.delete_cookie("session_token", path="/")
    return {"ok": True}

# ------------------------------------------------------------------
# Request Window
# ------------------------------------------------------------------
async def get_window_doc():
    doc = await db.settings.find_one({"key": "request_window"}, {"_id": 0})
    if not doc:
        doc = {"key": "request_window", "enabled": False, "opens_at": None, "closes_at": None}
        await db.settings.insert_one(doc)
        doc.pop("_id", None)
    return doc

def compute_window_state(doc):
    if not doc.get("enabled"):
        return {"is_open": False, "message": "Course change requests are currently closed.", "phase": "disabled"}
    now = now_utc()
    opens = doc.get("opens_at")
    closes = doc.get("closes_at")
    opens_dt = datetime.fromisoformat(opens) if opens else None
    closes_dt = datetime.fromisoformat(closes) if closes else None
    if opens_dt and opens_dt.tzinfo is None:
        opens_dt = opens_dt.replace(tzinfo=timezone.utc)
    if closes_dt and closes_dt.tzinfo is None:
        closes_dt = closes_dt.replace(tzinfo=timezone.utc)
    if opens_dt and now < opens_dt:
        return {"is_open": False, "message": "Course change requests are currently closed.", "phase": "before"}
    if closes_dt and now > closes_dt:
        return {"is_open": False, "message": "The request window has closed.", "phase": "after"}
    return {"is_open": True, "message": "Course change requests are open.", "phase": "open"}

@api_router.get("/window")
async def window_status(request: Request):
    await get_current_user(request)
    doc = await get_window_doc()
    state = compute_window_state(doc)
    return {**state, "opens_at": doc.get("opens_at"), "closes_at": doc.get("closes_at"), "enabled": doc.get("enabled")}

class WindowInput(BaseModel):
    enabled: bool
    opens_at: Optional[str] = None
    closes_at: Optional[str] = None

@api_router.put("/admin/window")
async def update_window(payload: WindowInput, admin=Depends(require_admin)):
    closes_at = payload.closes_at
    if payload.enabled and not closes_at:
        base = datetime.fromisoformat(payload.opens_at) if payload.opens_at else now_utc()
        if base.tzinfo is None:
            base = base.replace(tzinfo=timezone.utc)
        closes_at = iso(base + timedelta(hours=24))
    await db.settings.update_one({"key": "request_window"}, {"$set": {
        "enabled": payload.enabled,
        "opens_at": payload.opens_at,
        "closes_at": closes_at,
    }}, upsert=True)
    await audit("WINDOW_UPDATE", admin.get("email"), f"enabled={payload.enabled} opens={payload.opens_at} closes={closes_at}")
    doc = await get_window_doc()
    return {**compute_window_state(doc), "opens_at": doc.get("opens_at"), "closes_at": doc.get("closes_at"), "enabled": doc.get("enabled")}

# ------------------------------------------------------------------
# Master data lookups
# ------------------------------------------------------------------
async def build_course_section_maps():
    courses = await db.courses.find({}, {"_id": 0}).to_list(1000)
    sections = await db.sections.find({}, {"_id": 0}).to_list(2000)
    cmap = {c["course_id"]: c for c in courses}
    smap = {s["section_id"]: s for s in sections}
    return cmap, smap

async def enrich_enrollment(e, cmap, smap):
    c = cmap.get(e["course_id"], {})
    s = smap.get(e["section_id"], {})
    return {
        "course_id": e["course_id"],
        "section_id": e["section_id"],
        "course_code": c.get("course_code"),
        "course_name": c.get("course_name"),
        "credits": c.get("credits"),
        "area": c.get("area"),
        "section_name": s.get("section_name"),
        "day": s.get("day"),
        "time_slot": s.get("time_slot"),
        "mid_tag": s.get("mid_tag"),
    }

# ------------------------------------------------------------------
# Student endpoints
# ------------------------------------------------------------------
@api_router.get("/student/dashboard")
async def student_dashboard(student=Depends(require_student)):
    pgpid = student["pgpid"]
    cmap, smap = await build_course_section_maps()
    enrolls = await db.enrollments.find({"pgpid": pgpid}, {"_id": 0}).to_list(1000)
    courses = [await enrich_enrollment(e, cmap, smap) for e in enrolls]
    total_credits = round(sum((c.get("credits") or 0) for c in courses), 1)
    master = await db.students.find_one({"pgpid": pgpid}, {"_id": 0}) or {}
    program = master.get("program", "PGP")
    stex = bool(master.get("stex", False))
    rule_applies = (program == "PGP") and not stex
    credit_status = "ok"
    credit_message = f"Your total Term V credits: {total_credits}"
    if rule_applies:
        if total_credits < TERM_V_MIN_CREDITS:
            credit_status = "under"
            credit_message = f"You have {total_credits} credits. Term V requires between {TERM_V_MIN_CREDITS} and {TERM_V_MAX_CREDITS} credits — you are below the minimum."
        elif total_credits > TERM_V_MAX_CREDITS:
            credit_status = "over"
            credit_message = f"You have {total_credits} credits. Term V allows a maximum of {TERM_V_MAX_CREDITS} credits — you are above the limit."
        else:
            credit_message = f"You have {total_credits} credits — within the Term V range ({TERM_V_MIN_CREDITS}–{TERM_V_MAX_CREDITS})."
    doc = await get_window_doc()
    return {
        "name": student["name"],
        "pgpid": pgpid,
        "email": student["email"],
        "program": program,
        "stex": stex,
        "courses": courses,
        "total_credits": total_credits,
        "credit_status": credit_status,
        "credit_message": credit_message,
        "credit_min": TERM_V_MIN_CREDITS,
        "credit_max": TERM_V_MAX_CREDITS,
        "window": compute_window_state(doc),
    }

@api_router.get("/student/timetable")
async def student_timetable(student=Depends(require_student)):
    pgpid = student["pgpid"]
    cmap, smap = await build_course_section_maps()
    enrolls = await db.enrollments.find({"pgpid": pgpid}, {"_id": 0}).to_list(1000)
    entries = [await enrich_enrollment(e, cmap, smap) for e in enrolls]
    return {"entries": entries}

@api_router.get("/student/available-courses")
async def available_courses(student=Depends(require_student)):
    """Courses & sections WITHOUT any capacity/strength information (credits & schedule are allowed)."""
    courses = await db.courses.find({}, {"_id": 0}).to_list(1000)
    sections = await db.sections.find({}, {"_id": 0, "min_capacity": 0, "max_capacity": 0}).to_list(2000)
    by_course = {}
    for s in sections:
        by_course.setdefault(s["course_id"], []).append({
            "section_id": s["section_id"], "section_name": s["section_name"],
            "day": s.get("day"), "time_slot": s.get("time_slot"), "mid_tag": s.get("mid_tag"),
        })
    result = []
    for c in courses:
        secs = sorted(by_course.get(c["course_id"], []), key=lambda x: x["section_name"])
        result.append({"course_id": c["course_id"], "course_code": c["course_code"], "course_name": c["course_name"],
                       "credits": c.get("credits"), "area": c.get("area"), "sections": secs})
    return sorted(result, key=lambda x: x["course_name"])

@api_router.get("/student/timetable/all")
async def full_timetable(student=Depends(require_student)):
    cmap, smap = await build_course_section_maps()
    secs = await db.sections.find({}, {"_id": 0, "min_capacity": 0, "max_capacity": 0}).to_list(4000)
    out = []
    for s in secs:
        c = cmap.get(s["course_id"], {})
        out.append({"section_id": s["section_id"], "course_code": c.get("course_code"), "course_name": c.get("course_name"),
                    "section_name": s["section_name"], "day": s.get("day"), "time_slot": s.get("time_slot"),
                    "mid_tag": s.get("mid_tag"), "credits": c.get("credits")})
    return {"sections": out}

@api_router.get("/student/section/{section_id}/students")
async def section_students(section_id: str, student=Depends(require_student)):
    s = await db.sections.find_one({"section_id": section_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Section not found")
    cmap, _ = await build_course_section_maps()
    enr = await db.enrollments.find({"section_id": section_id}, {"_id": 0}).to_list(5000)
    names = {st["pgpid"]: st["name"] for st in await db.students.find({}, {"_id": 0}).to_list(10000)}
    roster = sorted([{"pgpid": e["pgpid"], "name": names.get(e["pgpid"], e["pgpid"])} for e in enr], key=lambda x: x["name"])
    c = cmap.get(s["course_id"], {})
    return {"course_name": c.get("course_name"), "course_code": c.get("course_code"), "section_name": s["section_name"], "students": roster}

@api_router.get("/student/lookup-partner")
async def lookup_partner(pgpid: str, student=Depends(require_student)):
    pgpid = pgpid.strip().upper()
    if pgpid == student["pgpid"]:
        raise HTTPException(status_code=400, detail="You cannot swap with yourself.")
    partner = await db.students.find_one({"pgpid": pgpid}, {"_id": 0})
    if not partner:
        raise HTTPException(status_code=404, detail="No student found with that PGPID.")
    cmap, smap = await build_course_section_maps()
    enrolls = await db.enrollments.find({"pgpid": pgpid}, {"_id": 0}).to_list(1000)
    courses = [await enrich_enrollment(e, cmap, smap) for e in enrolls]
    return {"pgpid": pgpid, "name": partner["name"], "courses": courses}

async def get_student_active_courses(pgpid: str) -> set:
    """course_ids referenced by student's non-terminal requests."""
    active = set()
    cursor = db.requests.find({"$or": [{"student_pgpid": pgpid}, {"swap.partner_pgpid": pgpid}], "status": {"$nin": list(TERMINAL_STATUSES)}}, {"_id": 0})
    async for req in cursor:
        for a in req.get("actions", []):
            if a.get("course_id"):
                active.add(a["course_id"])
        sw = req.get("swap")
        if sw:
            gv, gt = swap_pairs(sw)
            for reg in gv + gt:
                if reg.get("course_id"):
                    active.add(reg["course_id"])
    return active

class RequestInput(BaseModel):
    request_type: str
    comment: Optional[str] = None
    # add / drop / add_drop
    drop_course_id: Optional[str] = None
    drop_section_id: Optional[str] = None
    add_course_id: Optional[str] = None
    add_section_id: Optional[str] = None
    # swaps
    partner_pgpid: Optional[str] = None
    # course swap: initiator gives own course, wants partner's course
    give_course_id: Optional[str] = None
    give_section_id: Optional[str] = None
    want_course_id: Optional[str] = None
    want_section_id: Optional[str] = None
    give_section_ids: Optional[List[str]] = None
    want_section_ids: Optional[List[str]] = None
    # section swap: same course, own section -> requested section
    swap_course_id: Optional[str] = None
    my_section_id: Optional[str] = None
    requested_section_id: Optional[str] = None

async def ensure_window_open():
    doc = await get_window_doc()
    state = compute_window_state(doc)
    if not state["is_open"]:
        raise HTTPException(status_code=403, detail=state["message"])

def build_reg(section_id, cmap, smap):
    s = smap.get(section_id) or {}
    c = cmap.get(s.get("course_id"), {})
    return {"course_id": s.get("course_id"), "section_id": section_id,
            "course_code": c.get("course_code"),
            "course_name": c.get("course_name"), "section_name": s.get("section_name"),
            "credits": c.get("credits"), "day": s.get("day"), "time_slot": s.get("time_slot")}

def swap_pairs(sw):
    gives = sw.get("initiator_gives") or ([sw["initiator_current"]] if sw.get("initiator_current") else [])
    gets = sw.get("initiator_gets") or ([sw["initiator_requested"]] if sw.get("initiator_requested") else [])
    return gives, gets

def new_request_base(student, request_type, comment):    return {
        "request_id": f"R{uuid.uuid4().hex[:8].upper()}",
        "student_pgpid": student["pgpid"],
        "student_name": student["name"],
        "student_email": student["email"],
        "request_type": request_type,
        "comment": comment,
        "admin_comment": None,
        "actions": [],
        "swap": None,
        "credit_note": None,
        "clash_note": None,
        "created_at": iso(now_utc()),
        "updated_at": iso(now_utc()),
        "history": [],
    }

def add_history(req, status, by, note=""):
    req["status"] = status
    req["updated_at"] = iso(now_utc())
    req["history"].append({"status": status, "by": by, "note": note, "at": iso(now_utc())})

async def check_conflicts(pgpid, course_ids: set):
    active = await get_student_active_courses(pgpid)
    clash = active.intersection(course_ids)
    if clash:
        cmap, _ = await build_course_section_maps()
        names = [cmap.get(cid, {}).get("course_name", cid) for cid in clash]
        raise HTTPException(status_code=409, detail=f"You already have an active request involving: {', '.join(names)}. Please cancel it before submitting a conflicting request.")

async def compute_credit_note(pgpid, cmap, add_course_id=None, drop_course_id=None):
    enrolls = await db.enrollments.find({"pgpid": pgpid}, {"_id": 0}).to_list(1000)
    current = sum((cmap.get(e["course_id"], {}).get("credits") or 0) for e in enrolls)
    projected = current
    if drop_course_id:
        projected -= (cmap.get(drop_course_id, {}).get("credits") or 0)
    if add_course_id:
        projected += (cmap.get(add_course_id, {}).get("credits") or 0)
    projected = round(projected, 1)
    master = await db.students.find_one({"pgpid": pgpid}, {"_id": 0}) or {}
    if master.get("program", "PGP") == "PGP" and not master.get("stex", False):
        if projected > TERM_V_MAX_CREDITS:
            return f"Projected {projected} credits — above the Term V maximum of {TERM_V_MAX_CREDITS}."
        if projected < TERM_V_MIN_CREDITS:
            return f"Projected {projected} credits — below the Term V minimum of {TERM_V_MIN_CREDITS}."
    return None

QUOTA_EXCLUDE_STATUSES = {"REJECTED", "PARTNER_REJECTED", "CANCELLED"}
REQUEST_LIMITS = {"add": 1, "drop": 1, "course_swap": 2, "section_swap": 2}

async def get_request_quota(pgpid):
    reqs = await db.requests.find({"student_pgpid": pgpid, "status": {"$nin": list(QUOTA_EXCLUDE_STATUSES)}}, {"_id": 0, "request_type": 1}).to_list(1000)
    def n(t):
        return sum(1 for r in reqs if r["request_type"] == t)
    add_used = n("ADD") + n("ADD_DROP")
    drop_used = n("DROP") + n("ADD_DROP")
    return {
        "add_used": add_used, "add_limit": REQUEST_LIMITS["add"],
        "drop_used": drop_used, "drop_limit": REQUEST_LIMITS["drop"],
        "course_swap_used": n("COURSE_SWAP"), "course_swap_limit": REQUEST_LIMITS["course_swap"],
        "section_swap_used": n("SECTION_SWAP"), "section_swap_limit": REQUEST_LIMITS["section_swap"],
    }

async def compute_clash_note(pgpid, cmap, smap, new_section_id, exclude_course_id=None):
    ns = smap.get(new_section_id) or {}
    day = ns.get("day"); ts = ns.get("time_slot")
    if not day or day == "Not timetabled" or not ts or ts == "—":
        return None
    enrolls = await db.enrollments.find({"pgpid": pgpid}, {"_id": 0}).to_list(1000)
    clashes = []
    for e in enrolls:
        if exclude_course_id and e["course_id"] == exclude_course_id:
            continue
        s = smap.get(e["section_id"]) or {}
        if s.get("day") == day and s.get("time_slot") == ts:
            cn = cmap.get(e["course_id"], {}).get("course_name", "")
            clashes.append(f"{cn} (Sec {s.get('section_name')})")
    if clashes:
        return f"Time clash on {day} {ts} with: {', '.join(clashes)}."
    return None

def mid_norm(m):
    m = (m or "").upper()
    if "PRE" in m or m == "$":
        return "PRE"
    if "POST" in m or m == "#":
        return "POST"
    return "FULL"

def mids_overlap(a, b):
    na, nb = mid_norm(a), mid_norm(b)
    if na == "FULL" or nb == "FULL":
        return True
    return na == nb

def is_timetabled(s):
    return bool(s and s.get("day") and s.get("day") != "Not timetabled" and s.get("time_slot") and s.get("time_slot") != "—")

async def detect_clashes(pgpid):
    """Return real timetable clashes: >=2 enrolled sections in the same day+slot whose term-halves overlap."""
    cmap, smap = await build_course_section_maps()
    enrolls = await db.enrollments.find({"pgpid": pgpid}, {"_id": 0}).to_list(1000)
    regs = []
    for e in enrolls:
        s = smap.get(e["section_id"])
        if not is_timetabled(s):
            continue
        regs.append({**build_reg(e["section_id"], cmap, smap), "mid_tag": s.get("mid_tag")})
    groups = {}
    for i in range(len(regs)):
        for j in range(i + 1, len(regs)):
            a, b = regs[i], regs[j]
            if a["day"] == b["day"] and a["time_slot"] == b["time_slot"] and mids_overlap(a["mid_tag"], b["mid_tag"]):
                grp = groups.setdefault((a["day"], a["time_slot"]), {})
                grp[a["section_id"]] = a
                grp[b["section_id"]] = b
    return [{"day": d, "time_slot": t, "courses": list(g.values())} for (d, t), g in groups.items()]

@api_router.post("/student/requests")
async def submit_request(payload: RequestInput, student=Depends(require_student)):
    await ensure_window_open()
    rtype = payload.request_type
    if rtype not in REQUEST_TYPES:
        raise HTTPException(status_code=400, detail="Invalid request type")
    pgpid = student["pgpid"]
    cmap, smap = await build_course_section_maps()

    q = await get_request_quota(pgpid)
    if rtype == "ADD" and q["add_used"] >= q["add_limit"]:
        raise HTTPException(status_code=403, detail="You have reached the limit of 1 Add request.")
    if rtype == "DROP" and q["drop_used"] >= q["drop_limit"]:
        raise HTTPException(status_code=403, detail="You have reached the limit of 1 Drop request.")
    if rtype == "ADD_DROP" and (q["add_used"] >= q["add_limit"] or q["drop_used"] >= q["drop_limit"]):
        raise HTTPException(status_code=403, detail="You have reached your Add/Drop request limit (1 each).")
    if rtype == "COURSE_SWAP" and q["course_swap_used"] >= q["course_swap_limit"]:
        raise HTTPException(status_code=403, detail="You have reached the limit of 2 Course Swap requests.")
    if rtype == "SECTION_SWAP" and q["section_swap_used"] >= q["section_swap_limit"]:
        raise HTTPException(status_code=403, detail="You have reached the limit of 2 Section Swap requests.")

    async def owns(course_id, section_id):
        e = await db.enrollments.find_one({"pgpid": pgpid, "course_id": course_id, "section_id": section_id}, {"_id": 0})
        return e is not None

    def valid_section(course_id, section_id):
        s = smap.get(section_id)
        return s is not None and s["course_id"] == course_id

    req = new_request_base(student, rtype, payload.comment)

    if rtype == "ADD":
        if not payload.add_course_id or not payload.add_section_id or not valid_section(payload.add_course_id, payload.add_section_id):
            raise HTTPException(status_code=400, detail="Invalid course/section for add.")
        await check_conflicts(pgpid, {payload.add_course_id})
        req["actions"].append({"action": "ADD", "course_id": payload.add_course_id, "section_id": payload.add_section_id,
                               "course_name": cmap[payload.add_course_id]["course_name"], "section_name": smap[payload.add_section_id]["section_name"]})
        add_history(req, "SUBMITTED", pgpid, "Add request submitted")

    elif rtype == "DROP":
        if not payload.drop_course_id or not payload.drop_section_id or not await owns(payload.drop_course_id, payload.drop_section_id):
            raise HTTPException(status_code=400, detail="You are not enrolled in the selected course/section.")
        await check_conflicts(pgpid, {payload.drop_course_id})
        req["actions"].append({"action": "DROP", "course_id": payload.drop_course_id, "section_id": payload.drop_section_id,
                               "course_name": cmap[payload.drop_course_id]["course_name"], "section_name": smap[payload.drop_section_id]["section_name"]})
        add_history(req, "SUBMITTED", pgpid, "Drop request submitted")

    elif rtype == "ADD_DROP":
        if not payload.drop_course_id or not payload.drop_section_id or not await owns(payload.drop_course_id, payload.drop_section_id):
            raise HTTPException(status_code=400, detail="You are not enrolled in the course/section to drop.")
        if not payload.add_course_id or not payload.add_section_id or not valid_section(payload.add_course_id, payload.add_section_id):
            raise HTTPException(status_code=400, detail="Invalid course/section for add.")
        if payload.add_course_id == payload.drop_course_id:
            raise HTTPException(status_code=400, detail="Add and Drop cannot be the same course. Use Section Swap instead.")
        await check_conflicts(pgpid, {payload.add_course_id, payload.drop_course_id})
        req["actions"].append({"action": "DROP", "course_id": payload.drop_course_id, "section_id": payload.drop_section_id,
                               "course_name": cmap[payload.drop_course_id]["course_name"], "section_name": smap[payload.drop_section_id]["section_name"]})
        req["actions"].append({"action": "ADD", "course_id": payload.add_course_id, "section_id": payload.add_section_id,
                               "course_name": cmap[payload.add_course_id]["course_name"], "section_name": smap[payload.add_section_id]["section_name"]})
        add_history(req, "SUBMITTED", pgpid, "Add + Drop request submitted")

    elif rtype in ("COURSE_SWAP", "SECTION_SWAP"):
        partner_pgpid = (payload.partner_pgpid or "").strip().upper()
        if not partner_pgpid or partner_pgpid == pgpid:
            raise HTTPException(status_code=400, detail="Invalid swap partner.")
        partner = await db.students.find_one({"pgpid": partner_pgpid}, {"_id": 0})
        if not partner:
            raise HTTPException(status_code=404, detail="Swap partner PGPID not found.")

        if rtype == "COURSE_SWAP":
            give_ids = payload.give_section_ids or ([payload.give_section_id] if payload.give_section_id else [])
            want_ids = payload.want_section_ids or ([payload.want_section_id] if payload.want_section_id else [])
            if not give_ids or not want_ids:
                raise HTTPException(status_code=400, detail="Select at least one course to offer and one to receive.")
            gives = [build_reg(sid, cmap, smap) for sid in give_ids]
            gets = [build_reg(sid, cmap, smap) for sid in want_ids]
            for g in gives:
                if not g["course_id"] or not await owns(g["course_id"], g["section_id"]):
                    raise HTTPException(status_code=400, detail="You are not enrolled in a course/section you are offering.")
                if await db.enrollments.find_one({"pgpid": partner_pgpid, "course_id": g["course_id"]}, {"_id": 0}):
                    raise HTTPException(status_code=400, detail="Your swap partner already holds a course you are offering.")
            for g in gets:
                if not g["course_id"]:
                    raise HTTPException(status_code=400, detail="Invalid course selected to receive.")
                ph = await db.enrollments.find_one({"pgpid": partner_pgpid, "course_id": g["course_id"], "section_id": g["section_id"]}, {"_id": 0})
                if not ph:
                    raise HTTPException(status_code=400, detail="Your swap partner is not enrolled in a course/section you requested.")
                if await db.enrollments.find_one({"pgpid": pgpid, "course_id": g["course_id"]}, {"_id": 0}):
                    raise HTTPException(status_code=400, detail="You already have a course you are trying to receive.")
            give_courses = {g["course_id"] for g in gives}; get_courses = {g["course_id"] for g in gets}
            if give_courses & get_courses:
                raise HTTPException(status_code=400, detail="Courses you offer and receive must be different.")
            sg = round(sum(g["credits"] or 0 for g in gives), 1); sr = round(sum(g["credits"] or 0 for g in gets), 1)
            if sg != sr:
                raise HTTPException(status_code=400, detail=f"Total credits must match — you are offering {sg} and requesting {sr}.")
            await check_conflicts(pgpid, give_courses | get_courses)
            req["swap"] = {
                "kind": "COURSE", "partner_pgpid": partner_pgpid, "partner_name": partner["name"],
                "initiator_gives": gives, "initiator_gets": gets,
                "initiator_current": gives[0], "initiator_requested": gets[0],
                "partner_current": gets[0], "partner_requested": gives[0],
                "initiator_confirmed": True, "partner_confirmed": None,
            }
        else:  # SECTION_SWAP
            course_id = payload.swap_course_id
            if not await owns(course_id, payload.my_section_id):
                raise HTTPException(status_code=400, detail="You are not enrolled in the selected course/section.")
            if not valid_section(course_id, payload.requested_section_id) or payload.requested_section_id == payload.my_section_id:
                raise HTTPException(status_code=400, detail="Invalid requested section.")
            partner_has = await db.enrollments.find_one({"pgpid": partner_pgpid, "course_id": course_id, "section_id": payload.requested_section_id}, {"_id": 0})
            if not partner_has:
                raise HTTPException(status_code=400, detail="Swap partner is not in the section you requested.")
            await check_conflicts(pgpid, {course_id})
            g = build_reg(payload.my_section_id, cmap, smap); r2 = build_reg(payload.requested_section_id, cmap, smap)
            req["swap"] = {
                "kind": "SECTION", "partner_pgpid": partner_pgpid, "partner_name": partner["name"],
                "initiator_gives": [g], "initiator_gets": [r2],
                "initiator_current": g, "initiator_requested": r2,
                "partner_current": r2, "partner_requested": g,
                "initiator_confirmed": True, "partner_confirmed": None,
            }
        add_history(req, "AWAITING_PARTNER_CONFIRMATION", pgpid, f"Swap request sent to {partner_pgpid}")
        # notification for partner
        await db.notifications.insert_one({
            "notification_id": f"n_{uuid.uuid4().hex[:12]}",
            "pgpid": partner_pgpid,
            "type": "SWAP_REQUEST",
            "request_id": req["request_id"],
            "message": f"{pgpid} has requested a {req['swap']['kind'].lower()} swap with you.",
            "read": False,
            "created_at": iso(now_utc()),
        })

    if rtype in ("ADD", "DROP", "ADD_DROP"):
        req["credit_note"] = await compute_credit_note(
            pgpid, cmap,
            add_course_id=payload.add_course_id if rtype in ("ADD", "ADD_DROP") else None,
            drop_course_id=payload.drop_course_id if rtype in ("DROP", "ADD_DROP") else None,
        )

    if rtype in ("ADD", "ADD_DROP"):
        req["clash_note"] = await compute_clash_note(pgpid, cmap, smap, payload.add_section_id,
                                                     exclude_course_id=payload.drop_course_id if rtype == "ADD_DROP" else None)
    elif rtype == "COURSE_SWAP":
        notes = []
        for g in (req.get("swap", {}) or {}).get("initiator_gets", []):
            n = await compute_clash_note(pgpid, cmap, smap, g["section_id"])
            if n:
                notes.append(n)
        req["clash_note"] = " ".join(notes) if notes else None
    elif rtype == "SECTION_SWAP":
        req["clash_note"] = await compute_clash_note(pgpid, cmap, smap, payload.requested_section_id, exclude_course_id=payload.swap_course_id)

    await db.requests.insert_one({**req})
    await audit("SUBMIT", pgpid, f"submitted {rtype}", req["request_id"])
    req.pop("_id", None)
    return req

@api_router.get("/student/requests")
async def my_requests(student=Depends(require_student)):
    pgpid = student["pgpid"]
    reqs = await db.requests.find({"student_pgpid": pgpid}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return reqs

@api_router.get("/student/quota")
async def student_quota(student=Depends(require_student)):
    return await get_request_quota(student["pgpid"])

@api_router.get("/student/requests/{request_id}")
async def my_request_detail(request_id: str, student=Depends(require_student)):
    pgpid = student["pgpid"]
    req = await db.requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    involved = req["student_pgpid"] == pgpid or (req.get("swap") and req["swap"]["partner_pgpid"] == pgpid)
    if not involved:
        raise HTTPException(status_code=403, detail="Access denied")
    return req

@api_router.post("/student/requests/{request_id}/cancel")
async def cancel_request(request_id: str, student=Depends(require_student)):
    pgpid = student["pgpid"]
    req = await db.requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req or req["student_pgpid"] != pgpid:
        raise HTTPException(status_code=403, detail="Access denied")
    if req["request_type"] not in ("ADD", "DROP", "ADD_DROP"):
        raise HTTPException(status_code=403, detail="Swap requests cannot be withdrawn once submitted.")
    if not compute_window_state(await get_window_doc())["is_open"]:
        raise HTTPException(status_code=403, detail="The request window has closed — requests can no longer be withdrawn.")
    if req["status"] not in ("SUBMITTED", "UNDER_REVIEW"):
        raise HTTPException(status_code=400, detail="This request can no longer be withdrawn.")
    add_history(req, "CANCELLED", pgpid, "Withdrawn by student")
    await db.requests.update_one({"request_id": request_id}, {"$set": {"status": req["status"], "updated_at": req["updated_at"], "history": req["history"]}})
    await audit("WITHDRAW", pgpid, "withdrew request", request_id)
    return {"ok": True}

# ---- Notifications & swap response ----
@api_router.get("/student/notifications")
async def notifications(student=Depends(require_student)):
    pgpid = student["pgpid"]
    notes = await db.notifications.find({"pgpid": pgpid}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return notes

@api_router.post("/student/notifications/{notification_id}/read")
async def mark_read(notification_id: str, student=Depends(require_student)):
    await db.notifications.update_one({"notification_id": notification_id, "pgpid": student["pgpid"]}, {"$set": {"read": True}})
    return {"ok": True}

@api_router.get("/student/notifications/unread-count")
async def unread_count(student=Depends(require_student)):
    c = await db.notifications.count_documents({"pgpid": student["pgpid"], "read": False})
    return {"count": c}

@api_router.post("/student/notifications/read-all")
async def read_all(student=Depends(require_student)):
    await db.notifications.update_many({"pgpid": student["pgpid"], "read": False}, {"$set": {"read": True}})
    return {"ok": True}

@api_router.get("/student/pending-swaps")
async def pending_swaps(student=Depends(require_student)):
    pgpid = student["pgpid"]
    reqs = await db.requests.find({"swap.partner_pgpid": pgpid, "status": "AWAITING_PARTNER_CONFIRMATION"}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return reqs

class SwapResponseInput(BaseModel):
    action: str  # accept | reject

@api_router.post("/student/swaps/{request_id}/respond")
async def respond_swap(request_id: str, payload: SwapResponseInput, student=Depends(require_student)):
    pgpid = student["pgpid"]
    req = await db.requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req or not req.get("swap") or req["swap"]["partner_pgpid"] != pgpid:
        raise HTTPException(status_code=403, detail="This swap request is not addressed to you.")
    if req["status"] != "AWAITING_PARTNER_CONFIRMATION":
        raise HTTPException(status_code=400, detail="This swap can no longer be responded to.")
    swap = req["swap"]
    if payload.action == "accept":
        swap["partner_confirmed"] = True
        add_history(req, "BOTH_CONFIRMED", pgpid, "Partner accepted the swap")
        note = "Swap accepted — awaiting admin approval"
        await db.notifications.insert_one({
            "notification_id": f"n_{uuid.uuid4().hex[:12]}", "pgpid": req["student_pgpid"],
            "type": "SWAP_ACCEPTED", "request_id": request_id,
            "message": f"{pgpid} accepted your swap request. Awaiting admin approval.", "read": False, "created_at": iso(now_utc())})
    elif payload.action == "reject":
        swap["partner_confirmed"] = False
        add_history(req, "PARTNER_REJECTED", pgpid, "Partner rejected the swap")
        note = "Swap rejected by partner"
        await db.notifications.insert_one({
            "notification_id": f"n_{uuid.uuid4().hex[:12]}", "pgpid": req["student_pgpid"],
            "type": "SWAP_REJECTED", "request_id": request_id,
            "message": f"{pgpid} rejected your swap request.", "read": False, "created_at": iso(now_utc())})
    else:
        raise HTTPException(status_code=400, detail="Invalid action")
    await db.requests.update_one({"request_id": request_id}, {"$set": {"swap": swap, "status": req["status"], "updated_at": req["updated_at"], "history": req["history"]}})
    await audit("SWAP_RESPONSE", pgpid, note, request_id)
    return {"ok": True, "status": req["status"]}

# ---- Timetable clash resolution (always available, independent of request/trading windows) ----
@api_router.get("/student/clashes")
async def student_clashes(student=Depends(require_student)):
    return {"clashes": await detect_clashes(student["pgpid"])}

@api_router.get("/student/clash-options")
async def clash_options(drop_section_id: str, student=Depends(require_student)):
    """Eligible replacement courses/sections: not already held, valid, and clash-free after the drop. No capacity exposed."""
    pgpid = student["pgpid"]
    cmap, smap = await build_course_section_maps()
    enrolls = await db.enrollments.find({"pgpid": pgpid}, {"_id": 0}).to_list(1000)
    owned_courses = {e["course_id"] for e in enrolls}
    drop_course_id = None
    remaining = []
    for e in enrolls:
        if e["section_id"] == drop_section_id:
            drop_course_id = e["course_id"]
            continue
        s = smap.get(e["section_id"]) or {}
        if is_timetabled(s):
            remaining.append(s)
    if drop_course_id is None:
        raise HTTPException(status_code=400, detail="You are not enrolled in the selected section.")
    drop_credits = round(cmap.get(drop_course_id, {}).get("credits") or 0, 1)
    result = []
    for c in await db.courses.find({}, {"_id": 0}).to_list(1000):
        if c["course_id"] in owned_courses:
            continue
        secs = []
        for s in [x for x in smap.values() if x["course_id"] == c["course_id"]]:
            clash = False
            if is_timetabled(s):
                for r in remaining:
                    if r.get("day") == s.get("day") and r.get("time_slot") == s.get("time_slot") and mids_overlap(s.get("mid_tag"), r.get("mid_tag")):
                        clash = True
                        break
            if not clash:
                secs.append({"section_id": s["section_id"], "section_name": s["section_name"],
                             "day": s.get("day"), "time_slot": s.get("time_slot"), "mid_tag": s.get("mid_tag")})
        if secs:
            result.append({"course_id": c["course_id"], "course_code": c["course_code"], "course_name": c["course_name"],
                           "credits": c.get("credits"), "area": c.get("area"), "sections": sorted(secs, key=lambda x: x["section_name"])})
    return {"drop_course_id": drop_course_id, "required_credits": drop_credits,
            "courses": sorted(result, key=lambda x: x["course_name"])}

class ClashResolveInput(BaseModel):
    drop_section_id: str
    preferences: List[List[str]] = []

@api_router.post("/student/clash/resolve")
async def resolve_clash(payload: ClashResolveInput, student=Depends(require_student)):
    pgpid = student["pgpid"]
    cmap, smap = await build_course_section_maps()
    owned = await db.enrollments.find_one({"pgpid": pgpid, "section_id": payload.drop_section_id}, {"_id": 0})
    if not smap.get(payload.drop_section_id) or not owned:
        raise HTTPException(status_code=400, detail="You are not enrolled in the section you are trying to drop.")
    clashes = await detect_clashes(pgpid)
    target = next((cl for cl in clashes if any(cc["section_id"] == payload.drop_section_id for cc in cl["courses"])), None)
    if not target:
        raise HTTPException(status_code=400, detail="The selected section is not part of a timetable clash.")
    dup = await db.requests.find_one({"student_pgpid": pgpid, "request_type": "CLASH_RESOLUTION",
                                      "status": {"$nin": list(TERMINAL_STATUSES)},
                                      "clash.slot.day": target["day"], "clash.slot.time_slot": target["time_slot"]}, {"_id": 0})
    if dup:
        raise HTTPException(status_code=409, detail="You already have an active clash-resolution request for this slot.")
    if len(payload.preferences) < 2:
        raise HTTPException(status_code=400, detail="Please provide at least two replacement preferences.")
    drop_course_id = owned["course_id"]
    drop_credits = round(cmap.get(drop_course_id, {}).get("credits") or 0, 1)
    owned_courses = {e["course_id"] for e in await db.enrollments.find({"pgpid": pgpid}, {"_id": 0}).to_list(1000)}
    remaining = [smap.get(e["section_id"]) or {} for e in await db.enrollments.find({"pgpid": pgpid}, {"_id": 0}).to_list(1000) if e["section_id"] != payload.drop_section_id]
    pref_docs = []
    for idx, sec_ids in enumerate(payload.preferences, start=1):
        if not sec_ids:
            raise HTTPException(status_code=400, detail=f"Preference {idx} is empty.")
        items, seen, pref_slots, total = [], set(), [], 0.0
        for sid in sec_ids:
            s = smap.get(sid)
            if not s:
                raise HTTPException(status_code=400, detail="Invalid section selected.")
            cid = s["course_id"]
            cname = cmap.get(cid, {}).get("course_name", cid)
            if cid in owned_courses:
                raise HTTPException(status_code=400, detail=f"You already hold {cname}.")
            if cid in seen:
                raise HTTPException(status_code=400, detail="A preference cannot include the same course twice.")
            seen.add(cid)
            if is_timetabled(s):
                for r in remaining:
                    if r.get("day") == s.get("day") and r.get("time_slot") == s.get("time_slot") and mids_overlap(s.get("mid_tag"), r.get("mid_tag")):
                        raise HTTPException(status_code=400, detail=f"{cname} (Sec {s['section_name']}) clashes with your existing timetable.")
                for (pd, pt, pm) in pref_slots:
                    if pd == s.get("day") and pt == s.get("time_slot") and mids_overlap(s.get("mid_tag"), pm):
                        raise HTTPException(status_code=400, detail=f"Preference {idx}: the chosen courses clash with each other.")
                pref_slots.append((s.get("day"), s.get("time_slot"), s.get("mid_tag")))
            items.append({**build_reg(sid, cmap, smap), "mid_tag": s.get("mid_tag")})
            total += cmap.get(cid, {}).get("credits") or 0
        total = round(total, 1)
        if total != drop_credits:
            raise HTTPException(status_code=400, detail=f"Preference {idx} totals {total} credits but must equal the dropped course's {drop_credits} credits.")
        pref_docs.append({"rank": idx, "items": items, "total_credits": total})
    req = new_request_base(student, "CLASH_RESOLUTION", None)
    req["clash"] = {
        "slot": {"day": target["day"], "time_slot": target["time_slot"]},
        "drop": {**build_reg(payload.drop_section_id, cmap, smap)},
        "preferences": pref_docs,
        "approved_rank": None,
    }
    add_history(req, "SUBMITTED", pgpid, "Clash resolution request submitted")
    await db.requests.insert_one({**req})
    await audit("CLASH_SUBMIT", pgpid, f"clash resolution for {target['day']} {target['time_slot']}", req["request_id"])
    req.pop("_id", None)
    return req

# ------------------------------------------------------------------
# Admin endpoints
# ------------------------------------------------------------------
@api_router.get("/admin/dashboard")
async def admin_dashboard(admin=Depends(require_admin)):
    students = await db.students.count_documents({})
    courses = await db.courses.count_documents({})
    sections = await db.sections.count_documents({})
    all_reqs = await db.requests.find({}, {"_id": 0}).to_list(10000)
    def count_type(t):
        return sum(1 for r in all_reqs if r["request_type"] == t)
    awaiting_admin = sum(1 for r in all_reqs if r["status"] in ("SUBMITTED", "UNDER_REVIEW", "BOTH_CONFIRMED"))
    pending_swap_conf = sum(1 for r in all_reqs if r["status"] == "AWAITING_PARTNER_CONFIRMATION")
    return {
        "total_students": students,
        "total_courses": courses,
        "total_sections": sections,
        "total_requests": len(all_reqs),
        "add_requests": count_type("ADD"),
        "drop_requests": count_type("DROP"),
        "add_drop_requests": count_type("ADD_DROP"),
        "course_swaps": count_type("COURSE_SWAP"),
        "section_swaps": count_type("SECTION_SWAP"),
        "pending_swap_confirmations": pending_swap_conf,
        "awaiting_admin_review": awaiting_admin,
    }

@api_router.get("/admin/requests")
async def admin_requests(admin=Depends(require_admin)):
    reqs = await db.requests.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    return reqs

@api_router.get("/admin/requests/{request_id}")
async def admin_request_detail(request_id: str, admin=Depends(require_admin)):
    req = await db.requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Not found")
    return req

class DecisionInput(BaseModel):
    decision: str  # approve | reject | executed | under_review
    comment: Optional[str] = None

async def apply_execution(req):
    """Apply the actual enrollment change when a request is marked EXECUTED."""
    pgpid = req["student_pgpid"]
    if req.get("swap"):
        sw = req["swap"]; partner = sw["partner_pgpid"]
        gives, gets = swap_pairs(sw)
        for reg in gives:
            await db.enrollments.delete_one({"pgpid": pgpid, "course_id": reg["course_id"], "section_id": reg["section_id"]})
        for reg in gets:
            await db.enrollments.delete_one({"pgpid": partner, "course_id": reg["course_id"], "section_id": reg["section_id"]})
        for reg in gets:
            await db.enrollments.update_one({"pgpid": pgpid, "course_id": reg["course_id"]}, {"$set": {"section_id": reg["section_id"]}}, upsert=True)
        for reg in gives:
            await db.enrollments.update_one({"pgpid": partner, "course_id": reg["course_id"]}, {"$set": {"section_id": reg["section_id"]}}, upsert=True)
    else:
        for a in req.get("actions", []):
            if a["action"] == "DROP":
                await db.enrollments.delete_one({"pgpid": pgpid, "course_id": a["course_id"], "section_id": a["section_id"]})
        for a in req.get("actions", []):
            if a["action"] == "ADD":
                existing = await db.enrollments.find_one({"pgpid": pgpid, "course_id": a["course_id"]}, {"_id": 0})
                if existing:
                    await db.enrollments.update_one({"pgpid": pgpid, "course_id": a["course_id"]}, {"$set": {"section_id": a["section_id"]}})
                else:
                    await db.enrollments.insert_one({"pgpid": pgpid, "course_id": a["course_id"], "section_id": a["section_id"]})

@api_router.post("/admin/requests/{request_id}/decision")
async def admin_decision(request_id: str, payload: DecisionInput, admin=Depends(require_admin)):
    req = await db.requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Not found")
    is_swap = req.get("swap") is not None
    status = req["status"]
    d = payload.decision
    if is_swap and d == "approve" and status != "BOTH_CONFIRMED":
        raise HTTPException(status_code=400, detail="Swap can only be approved after both students have confirmed.")
    if d == "under_review":
        add_history(req, "UNDER_REVIEW", admin["email"], payload.comment or "Marked under review")
    elif d == "approve":
        add_history(req, "APPROVED_PENDING_EXECUTION", admin["email"], payload.comment or "Approved — pending execution")
    elif d == "reject":
        add_history(req, "REJECTED", admin["email"], payload.comment or "Rejected by admin")
    elif d == "executed":
        if status != "APPROVED_PENDING_EXECUTION":
            raise HTTPException(status_code=400, detail="Only approved requests can be marked executed.")
        await apply_execution(req)
        add_history(req, "EXECUTED", admin["email"], payload.comment or "Executed — enrollment updated")
    else:
        raise HTTPException(status_code=400, detail="Invalid decision")
    req["admin_comment"] = payload.comment
    await db.requests.update_one({"request_id": request_id}, {"$set": {"status": req["status"], "admin_comment": req["admin_comment"], "updated_at": req["updated_at"], "history": req["history"]}})
    # notify involved students
    targets = [req["student_pgpid"]]
    if is_swap:
        targets.append(req["swap"]["partner_pgpid"])
    for t in targets:
        await db.notifications.insert_one({"notification_id": f"n_{uuid.uuid4().hex[:12]}", "pgpid": t, "type": "ADMIN_DECISION", "request_id": request_id, "message": f"Request {request_id} is now {req['status'].replace('_',' ').title()}.", "read": False, "created_at": iso(now_utc())})
    await audit("ADMIN_DECISION", admin["email"], f"{d} -> {req['status']}", request_id)
    return {"ok": True, "status": req["status"]}

# ---- Admin: timetable clash resolution queue ----
async def notify_student(pgpid, request_id, message, ntype="ADMIN_DECISION"):
    await db.notifications.insert_one({"notification_id": f"n_{uuid.uuid4().hex[:12]}", "pgpid": pgpid,
                                       "type": ntype, "request_id": request_id, "message": message,
                                       "read": False, "created_at": iso(now_utc())})

@api_router.get("/admin/clashes")
async def admin_clashes(admin=Depends(require_admin)):
    reqs = await db.requests.find({"request_type": "CLASH_RESOLUTION"}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return reqs

@api_router.get("/admin/clash-tracker")
async def admin_clash_tracker(admin=Depends(require_admin)):
    """One-place overview of every timetable clash and whether it's resolved."""
    from collections import defaultdict
    cmap, smap = await build_course_section_maps()
    enrolls = await db.enrollments.find({}, {"_id": 0}).to_list(100000)
    students = {s["pgpid"]: s["name"] for s in await db.students.find({}, {"_id": 0}).to_list(20000)}
    perstu = defaultdict(lambda: defaultdict(list))
    for e in enrolls:
        s = smap.get(e["section_id"])
        if not is_timetabled(s):
            continue
        perstu[e["pgpid"]][(s["day"], s["time_slot"])].append(s)
    current = {}
    for pgpid, slots in perstu.items():
        for (day, ts), secs in slots.items():
            clashing = {}
            for i in range(len(secs)):
                for j in range(i + 1, len(secs)):
                    if mids_overlap(secs[i].get("mid_tag"), secs[j].get("mid_tag")):
                        clashing[secs[i]["section_id"]] = secs[i]
                        clashing[secs[j]["section_id"]] = secs[j]
            if clashing:
                current[(pgpid, day, ts)] = [f"{cmap.get(s['course_id'], {}).get('course_name')} (Sec {s['section_name']})" for s in clashing.values()]
    reqs = await db.requests.find({"request_type": "CLASH_RESOLUTION"}, {"_id": 0}).sort("created_at", 1).to_list(5000)
    reqmap = {}
    for r in reqs:
        reqmap[(r["student_pgpid"], r["clash"]["slot"]["day"], r["clash"]["slot"]["time_slot"])] = r
    entries = []
    for k in set(current.keys()) | set(reqmap.keys()):
        pgpid, day, ts = k
        req = reqmap.get(k)
        if k in current:
            status = "IN_PROGRESS" if (req and req["status"] not in ("EXECUTED", "REJECTED")) else "UNRESOLVED"
            courses = current[k]
        else:
            if not req:
                continue
            status = "RESOLVED"
            courses = [f"Dropped {req['clash']['drop']['course_name']} (Sec {req['clash']['drop']['section_name']})"]
        entries.append({
            "pgpid": pgpid, "name": students.get(pgpid, pgpid), "day": day, "time_slot": ts,
            "courses": courses, "status": status,
            "request_id": req["request_id"] if req else None,
            "request_status": req["status"] if req else None,
        })
    order = {"UNRESOLVED": 0, "IN_PROGRESS": 1, "RESOLVED": 2}
    entries.sort(key=lambda x: (order.get(x["status"], 9), x["pgpid"]))
    summary = {
        "total": len(entries),
        "unresolved": sum(1 for e in entries if e["status"] == "UNRESOLVED"),
        "in_progress": sum(1 for e in entries if e["status"] == "IN_PROGRESS"),
        "resolved": sum(1 for e in entries if e["status"] == "RESOLVED"),
    }
    return {"summary": summary, "entries": entries}

class ClashApproveInput(BaseModel):
    preference_rank: int
    comment: Optional[str] = None

class ClashCommentInput(BaseModel):
    comment: Optional[str] = None

@api_router.post("/admin/clashes/{request_id}/approve")
async def admin_clash_approve(request_id: str, payload: ClashApproveInput, admin=Depends(require_admin)):
    req = await db.requests.find_one({"request_id": request_id, "request_type": "CLASH_RESOLUTION"}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Clash request not found")
    if req["status"] not in ("SUBMITTED", "UNDER_REVIEW"):
        raise HTTPException(status_code=400, detail="This clash request can no longer be approved.")
    if not any(p["rank"] == payload.preference_rank for p in req["clash"]["preferences"]):
        raise HTTPException(status_code=400, detail="Invalid preference selected.")
    req["clash"]["approved_rank"] = payload.preference_rank
    req["admin_comment"] = payload.comment
    add_history(req, "APPROVED_PENDING_EXECUTION", admin["email"], payload.comment or f"Approved preference {payload.preference_rank} — pending execution")
    await db.requests.update_one({"request_id": request_id}, {"$set": {"status": req["status"], "clash": req["clash"], "admin_comment": req["admin_comment"], "updated_at": req["updated_at"], "history": req["history"]}})
    await notify_student(req["student_pgpid"], request_id, f"Your clash-resolution request {request_id} was approved (preference {payload.preference_rank}) — pending execution.")
    await audit("CLASH_APPROVE", admin["email"], f"approved preference {payload.preference_rank}", request_id)
    return {"ok": True, "status": req["status"]}

@api_router.post("/admin/clashes/{request_id}/reject")
async def admin_clash_reject(request_id: str, payload: ClashCommentInput, admin=Depends(require_admin)):
    req = await db.requests.find_one({"request_id": request_id, "request_type": "CLASH_RESOLUTION"}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Clash request not found")
    if req["status"] in ("EXECUTED", "REJECTED"):
        raise HTTPException(status_code=400, detail="This clash request is already closed.")
    req["admin_comment"] = payload.comment
    add_history(req, "REJECTED", admin["email"], payload.comment or "Rejected by admin")
    await db.requests.update_one({"request_id": request_id}, {"$set": {"status": req["status"], "admin_comment": req["admin_comment"], "updated_at": req["updated_at"], "history": req["history"]}})
    await notify_student(req["student_pgpid"], request_id, f"Your clash-resolution request {request_id} was rejected.")
    await audit("CLASH_REJECT", admin["email"], "rejected clash resolution", request_id)
    return {"ok": True, "status": req["status"]}

@api_router.post("/admin/clashes/{request_id}/execute")
async def admin_clash_execute(request_id: str, payload: ClashCommentInput, admin=Depends(require_admin)):
    req = await db.requests.find_one({"request_id": request_id, "request_type": "CLASH_RESOLUTION"}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Clash request not found")
    if req["status"] != "APPROVED_PENDING_EXECUTION":
        raise HTTPException(status_code=400, detail="Only approved clash requests can be executed.")
    rank = req["clash"].get("approved_rank")
    pref = next((p for p in req["clash"]["preferences"] if p["rank"] == rank), None)
    if not pref:
        raise HTTPException(status_code=400, detail="Approved preference not found.")
    pgpid = req["student_pgpid"]
    drop = req["clash"]["drop"]
    await db.enrollments.delete_one({"pgpid": pgpid, "course_id": drop["course_id"], "section_id": drop["section_id"]})
    for item in pref["items"]:
        existing = await db.enrollments.find_one({"pgpid": pgpid, "course_id": item["course_id"]}, {"_id": 0})
        if existing:
            await db.enrollments.update_one({"pgpid": pgpid, "course_id": item["course_id"]}, {"$set": {"section_id": item["section_id"]}})
        else:
            await db.enrollments.insert_one({"pgpid": pgpid, "course_id": item["course_id"], "section_id": item["section_id"]})
    req["admin_comment"] = payload.comment
    add_history(req, "EXECUTED", admin["email"], payload.comment or "Executed — enrollment updated, clash resolved")
    await db.requests.update_one({"request_id": request_id}, {"$set": {"status": req["status"], "admin_comment": req["admin_comment"], "updated_at": req["updated_at"], "history": req["history"]}})
    await notify_student(req["student_pgpid"], request_id, f"Your clash-resolution request {request_id} was executed — your timetable is updated.")
    await audit("CLASH_EXECUTE", admin["email"], f"executed preference {rank}", request_id)
    return {"ok": True, "status": req["status"]}

@api_router.get("/admin/students")
async def admin_students(admin=Depends(require_admin)):
    return await db.students.find({}, {"_id": 0}).sort("pgpid", 1).to_list(5000)

@api_router.get("/admin/courses")
async def admin_courses(admin=Depends(require_admin)):
    return await db.courses.find({}, {"_id": 0}).sort("course_code", 1).to_list(2000)

@api_router.get("/admin/sections")
async def admin_sections(admin=Depends(require_admin)):
    cmap, _ = await build_course_section_maps()
    secs = await db.sections.find({}, {"_id": 0}).to_list(4000)
    for s in secs:
        s["course_name"] = cmap.get(s["course_id"], {}).get("course_name")
        s["course_code"] = cmap.get(s["course_id"], {}).get("course_code")
    return sorted(secs, key=lambda x: (x.get("course_code") or "", x["section_name"]))

async def compute_capacity():
    cmap, smap = await build_course_section_maps()
    sections = await db.sections.find({}, {"_id": 0}).to_list(4000)
    # current enrollment per section
    enrolls = await db.enrollments.find({}, {"_id": 0}).to_list(50000)
    current = {}
    for e in enrolls:
        current[e["section_id"]] = current.get(e["section_id"], 0) + 1
    # pending adds/drops per section from non-terminal requests
    adds = {}
    drops = {}
    cursor = db.requests.find({"status": {"$nin": list(TERMINAL_STATUSES)}}, {"_id": 0})
    async for r in cursor:
        for a in r.get("actions", []):
            if a["action"] == "ADD":
                adds[a["section_id"]] = adds.get(a["section_id"], 0) + 1
            elif a["action"] == "DROP":
                drops[a["section_id"]] = drops.get(a["section_id"], 0) + 1
        sw = r.get("swap")
        if sw:
            gives, gets = swap_pairs(sw)
            for reg in gives:
                drops[reg["section_id"]] = drops.get(reg["section_id"], 0) + 1
                adds[reg["section_id"]] = adds.get(reg["section_id"], 0) + 1
            for reg in gets:
                adds[reg["section_id"]] = adds.get(reg["section_id"], 0) + 1
                drops[reg["section_id"]] = drops.get(reg["section_id"], 0) + 1
    rows = []
    for s in sorted(sections, key=lambda x: (cmap.get(x["course_id"], {}).get("course_code") or "", x["section_name"])):
        sid = s["section_id"]
        cur = current.get(sid, 0)
        pa = adds.get(sid, 0)
        pd = drops.get(sid, 0)
        rows.append({
            "section_id": sid,
            "course_code": cmap.get(s["course_id"], {}).get("course_code"),
            "course_name": cmap.get(s["course_id"], {}).get("course_name"),
            "section_name": s["section_name"],
            "current": cur,
            "min_capacity": s.get("min_capacity"),
            "max_capacity": s.get("max_capacity"),
            "pending_adds": pa,
            "pending_drops": pd,
            "net_change": pa - pd,
            "projected": cur + pa - pd,
        })
    return rows

@api_router.get("/admin/capacity")
async def admin_capacity(admin=Depends(require_admin)):
    return await compute_capacity()

@api_router.get("/admin/audit")
async def admin_audit(admin=Depends(require_admin)):
    return await db.audit_logs.find({}, {"_id": 0}).sort("at", -1).to_list(2000)

# ---- Export ----
@api_router.get("/admin/export")
async def admin_export(admin=Depends(require_admin)):
    reqs = await db.requests.find({}, {"_id": 0}).sort("created_at", 1).to_list(20000)
    rows = []
    for r in reqs:
        actions = {a["action"]: a for a in r.get("actions", [])}
        sw = r.get("swap") or {}
        drop = actions.get("DROP", {})
        add = actions.get("ADD", {})
        current_course = current_section = ""
        add_course = add_section = ""
        drop_course = drop_section = ""
        if sw:
            current_course = sw.get("initiator_current", {}).get("course_name", "")
            current_section = sw.get("initiator_current", {}).get("section_name", "")
            add_course = sw.get("initiator_requested", {}).get("course_name", "")
            add_section = sw.get("initiator_requested", {}).get("section_name", "")
        rows.append({
            "Request ID": r["request_id"],
            "PGPID": r["student_pgpid"],
            "Student Name": r["student_name"],
            "Student Email": r["student_email"],
            "Request Type": r["request_type"],
            "Current Course": current_course,
            "Current Section": current_section,
            "Course to Drop": drop.get("course_name", ""),
            "Section to Drop": drop.get("section_name", ""),
            "Course to Add": add.get("course_name", "") or add_course,
            "Preferred Add Section": add.get("section_name", "") or add_section,
            "Swap Partner PGPID": sw.get("partner_pgpid", ""),
            "Swap Partner Name": sw.get("partner_name", ""),
            "Partner Confirmation": ("" if not sw else ("Accepted" if sw.get("partner_confirmed") is True else ("Rejected" if sw.get("partner_confirmed") is False else "Pending"))),
            "Request Status": r["status"],
            "Student Comment": r.get("comment", ""),
            "Admin Comment": r.get("admin_comment", ""),
            "Submitted At": r["created_at"],
            "Updated At": r["updated_at"],
        })
    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Requests")
    buf.seek(0)
    await audit("EXPORT", admin["email"], f"exported {len(rows)} requests")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=course_change_requests.xlsx"})

@api_router.get("/admin/export/executed")
async def admin_export_executed(admin=Depends(require_admin)):
    reqs = await db.requests.find({"status": "EXECUTED"}, {"_id": 0}).sort("updated_at", 1).to_list(20000)
    rows = []
    for r in reqs:
        sw = r.get("swap")
        if sw:
            rows.append({"PGPID": r["student_pgpid"], "Student Name": r["student_name"], "Change": "SWAP OUT",
                         "Course": sw["initiator_current"]["course_name"], "From Section": sw["initiator_current"]["section_name"], "To Section": "", "Request ID": r["request_id"], "Executed At": r["updated_at"]})
            rows.append({"PGPID": r["student_pgpid"], "Student Name": r["student_name"], "Change": "SWAP IN",
                         "Course": sw["initiator_requested"]["course_name"], "From Section": "", "To Section": sw["initiator_requested"]["section_name"], "Request ID": r["request_id"], "Executed At": r["updated_at"]})
            rows.append({"PGPID": sw["partner_pgpid"], "Student Name": sw["partner_name"], "Change": "SWAP OUT",
                         "Course": sw["partner_current"]["course_name"], "From Section": sw["partner_current"]["section_name"], "To Section": "", "Request ID": r["request_id"], "Executed At": r["updated_at"]})
            rows.append({"PGPID": sw["partner_pgpid"], "Student Name": sw["partner_name"], "Change": "SWAP IN",
                         "Course": sw["partner_requested"]["course_name"], "From Section": "", "To Section": sw["partner_requested"]["section_name"], "Request ID": r["request_id"], "Executed At": r["updated_at"]})
        else:
            for a in r.get("actions", []):
                rows.append({"PGPID": r["student_pgpid"], "Student Name": r["student_name"], "Change": a["action"],
                             "Course": a["course_name"], "From Section": a["section_name"] if a["action"] == "DROP" else "", "To Section": a["section_name"] if a["action"] == "ADD" else "", "Request ID": r["request_id"], "Executed At": r["updated_at"]})
    df = pd.DataFrame(rows, columns=["PGPID", "Student Name", "Change", "Course", "From Section", "To Section", "Request ID", "Executed At"])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Executed Changes")
    buf.seek(0)
    await audit("EXPORT_EXECUTED", admin["email"], f"exported {len(rows)} executed change rows")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=executed_changes.xlsx"})

# ---- Master data import ----
def read_upload_to_df(content: bytes, filename: str) -> pd.DataFrame:
    if filename.lower().endswith(".csv"):
        return pd.read_csv(io.BytesIO(content), dtype=str).fillna("")
    return pd.read_excel(io.BytesIO(content), dtype=str).fillna("")

def norm_cols(df):
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df

async def validate_import(kind: str, df: pd.DataFrame):
    df = norm_cols(df)
    rows = []
    valid_count = 0
    error_count = 0
    existing_students = {s["pgpid"]: s for s in await db.students.find({}, {"_id": 0}).to_list(10000)}
    courses = await db.courses.find({}, {"_id": 0}).to_list(2000)
    course_by_code = {c["course_code"].upper(): c for c in courses}
    sections = await db.sections.find({}, {"_id": 0}).to_list(4000)
    section_lookup = {(s["course_id"], s["section_name"].upper()): s for s in sections}

    for _, r in df.iterrows():
        errors = []
        data = {}
        if kind == "students":
            pgpid = (r.get("pgpid") or r.get("pgp id") or "").strip().upper()
            name = (r.get("name") or "").strip()
            email = (r.get("email") or "").strip().lower()
            if not pgpid: errors.append("Missing PGPID")
            if not name: errors.append("Missing Name")
            if not email or "@" not in email: errors.append("Invalid Email")
            data = {"pgpid": pgpid, "name": name, "email": email}
        elif kind == "courses":
            code = (r.get("course_code") or r.get("course code") or r.get("code") or "").strip().upper()
            name = (r.get("course_name") or r.get("course name") or r.get("name") or "").strip()
            if not code: errors.append("Missing Course Code")
            if not name: errors.append("Missing Course Name")
            data = {"course_code": code, "course_name": name}
        elif kind == "sections":
            code = (r.get("course_code") or r.get("course code") or r.get("code") or "").strip().upper()
            section = (r.get("section") or r.get("section_name") or "").strip().upper()
            mn = (r.get("min") or r.get("min_capacity") or r.get("minimum") or "").strip()
            mx = (r.get("max") or r.get("max_capacity") or r.get("maximum") or "").strip()
            if code not in course_by_code: errors.append(f"Course '{code}' not found")
            if not section: errors.append("Missing Section")
            try:
                mn_i = int(float(mn)); mx_i = int(float(mx))
                if mn_i > mx_i: errors.append("Min > Max")
            except Exception:
                errors.append("Invalid Min/Max"); mn_i = 0; mx_i = 0
            data = {"course_code": code, "section_name": section, "min_capacity": mn_i, "max_capacity": mx_i}
        elif kind == "enrollments":
            pgpid = (r.get("pgpid") or r.get("pgp id") or "").strip().upper()
            code = (r.get("course_code") or r.get("course code") or r.get("code") or "").strip().upper()
            section = (r.get("section") or r.get("section_name") or "").strip().upper()
            if pgpid not in existing_students: errors.append(f"Student '{pgpid}' not found")
            course = course_by_code.get(code)
            if not course:
                errors.append(f"Course '{code}' not found")
            elif (course["course_id"], section) not in section_lookup:
                errors.append(f"Section '{section}' not found for {code}")
            data = {"pgpid": pgpid, "course_code": code, "section_name": section}
        else:
            raise HTTPException(status_code=400, detail="Invalid import kind")

        valid = len(errors) == 0
        if valid: valid_count += 1
        else: error_count += 1
        rows.append({"data": data, "valid": valid, "errors": errors})
    return rows, valid_count, error_count

@api_router.post("/admin/import/preview")
async def import_preview(kind: str = Form(...), file: UploadFile = File(...), admin=Depends(require_admin)):
    content = await file.read()
    df = read_upload_to_df(content, file.filename)
    rows, valid_count, error_count = await validate_import(kind, df)
    token = f"imp_{uuid.uuid4().hex[:12]}"
    await db.import_staging.insert_one({"token": token, "kind": kind, "rows": rows, "created_at": iso(now_utc())})
    return {"token": token, "kind": kind, "total": len(rows), "valid": valid_count, "errors": error_count,
            "columns": list(df.columns), "rows": rows[:200]}

class CommitInput(BaseModel):
    token: str

@api_router.post("/admin/import/commit")
async def import_commit(payload: CommitInput, admin=Depends(require_admin)):
    staging = await db.import_staging.find_one({"token": payload.token}, {"_id": 0})
    if not staging:
        raise HTTPException(status_code=404, detail="Import session expired. Please re-upload.")
    kind = staging["kind"]
    valid_rows = [r["data"] for r in staging["rows"] if r["valid"]]
    inserted = 0
    course_by_code = {c["course_code"].upper(): c for c in await db.courses.find({}, {"_id": 0}).to_list(2000)}

    for d in valid_rows:
        if kind == "students":
            await db.students.update_one({"pgpid": d["pgpid"]}, {"$set": d}, upsert=True)
            inserted += 1
        elif kind == "courses":
            existing = course_by_code.get(d["course_code"])
            if existing:
                await db.courses.update_one({"course_id": existing["course_id"]}, {"$set": {"course_name": d["course_name"]}})
            else:
                cid = f"course_{uuid.uuid4().hex[:10]}"
                await db.courses.insert_one({"course_id": cid, "course_code": d["course_code"], "course_name": d["course_name"]})
                course_by_code[d["course_code"]] = {"course_id": cid, **d}
            inserted += 1
    # re-fetch maps for sections/enrollments
    if kind == "sections":
        course_by_code = {c["course_code"].upper(): c for c in await db.courses.find({}, {"_id": 0}).to_list(2000)}
        for d in valid_rows:
            course = course_by_code.get(d["course_code"])
            if not course: continue
            existing = await db.sections.find_one({"course_id": course["course_id"], "section_name": d["section_name"]}, {"_id": 0})
            payload_s = {"course_id": course["course_id"], "section_name": d["section_name"], "min_capacity": d["min_capacity"], "max_capacity": d["max_capacity"]}
            if existing:
                await db.sections.update_one({"section_id": existing["section_id"]}, {"$set": payload_s})
            else:
                sid = f"section_{uuid.uuid4().hex[:10]}"
                await db.sections.insert_one({"section_id": sid, **payload_s})
            inserted += 1
    if kind == "enrollments":
        course_by_code = {c["course_code"].upper(): c for c in await db.courses.find({}, {"_id": 0}).to_list(2000)}
        for d in valid_rows:
            course = course_by_code.get(d["course_code"])
            if not course: continue
            section = await db.sections.find_one({"course_id": course["course_id"], "section_name": d["section_name"]}, {"_id": 0})
            if not section: continue
            await db.enrollments.update_one({"pgpid": d["pgpid"], "course_id": course["course_id"]},
                                            {"$set": {"pgpid": d["pgpid"], "course_id": course["course_id"], "section_id": section["section_id"]}}, upsert=True)
            inserted += 1

    await db.import_staging.delete_one({"token": payload.token})
    await audit("IMPORT", admin["email"], f"imported {inserted} {kind} records")
    return {"ok": True, "inserted": inserted, "kind": kind}

class SectionLimits(BaseModel):
    min_capacity: Optional[int] = None
    max_capacity: Optional[int] = None

@api_router.put("/admin/sections/{section_id}")
async def update_section_limits(section_id: str, payload: SectionLimits, admin=Depends(require_admin)):
    s = await db.sections.find_one({"section_id": section_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Section not found")
    if payload.min_capacity is not None and payload.max_capacity is not None and payload.min_capacity > payload.max_capacity:
        raise HTTPException(status_code=400, detail="Minimum capacity cannot be greater than maximum capacity.")
    await db.sections.update_one({"section_id": section_id}, {"$set": {"min_capacity": payload.min_capacity, "max_capacity": payload.max_capacity}})
    await audit("SECTION_LIMITS", admin["email"], f"set limits for {section_id}: min={payload.min_capacity} max={payload.max_capacity}")
    return {"ok": True}

@api_router.get("/admin/feasibility")
async def feasibility(admin=Depends(require_admin)):
    rows = await compute_capacity()
    summary = {"OK": 0, "OVER": 0, "UNDER": 0, "NO_LIMIT": 0}
    for r in rows:
        mx, mn, proj = r["max_capacity"], r["min_capacity"], r["projected"]
        if mx is None and mn is None:
            flag = "NO_LIMIT"
        elif mx is not None and proj > mx:
            flag = "OVER"
        elif mn is not None and proj < mn:
            flag = "UNDER"
        else:
            flag = "OK"
        r["flag"] = flag
        summary[flag] += 1
    return {"sections": rows, "summary": summary}

# ---- Term V consolidated workbook import ----
def _clean(v):
    return str(v).strip() if v is not None else ""

@api_router.post("/admin/import/termv")
async def import_termv(file: UploadFile = File(...), admin=Depends(require_admin)):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read the workbook. Please upload the Term V .xlsx file.")

    # 1) Course metadata (credits, area) from "Courses & Sections"
    course_meta = {}
    if "Courses & Sections" in wb.sheetnames:
        ws = wb["Courses & Sections"]
        for r in ws.iter_rows(min_row=4, values_only=True):
            if r and _clean(r[0]) == "Course":
                short = _clean(r[1])
                if not short:
                    continue
                course_meta[short] = {
                    "course_name": _clean(r[3]),
                    "long_code": _clean(r[2]),
                    "area": _clean(r[5]),
                    "credits": float(r[6]) if r[6] is not None else 0.0,
                }

    if "Students by Section" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Missing 'Students by Section' sheet.")

    ws = wb["Students by Section"]
    students = {}       # roll_no -> student dict
    courses = {}        # short_code -> course dict
    sections = {}       # (short_code, section_name) -> section dict
    enrollments = []    # (roll_no, short_code, section_name)

    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or not _clean(r[0]):
            continue
        short = _clean(r[0]); cname = _clean(r[1]); section = _clean(r[2]); mid = _clean(r[3])
        day = _clean(r[4]); ts = _clean(r[5]); reg = _clean(r[7]); roll = _clean(r[8])
        name = _clean(r[9]); email = _clean(r[10]).lower(); prog = _clean(r[11])
        if not roll:
            continue
        students[roll] = {"pgpid": roll.upper(), "registration_id": reg, "name": name,
                          "email": email or f"{roll.lower()}@iiml.ac.in", "program": prog, "stex": False}
        meta = course_meta.get(short, {})
        courses.setdefault(short, {"course_code": short, "course_name": cname or meta.get("course_name", short),
                                   "credits": meta.get("credits", 0.0), "area": meta.get("area", ""), "long_code": meta.get("long_code", "")})
        skey = (short, section)
        sections.setdefault(skey, {"section_name": section, "day": day, "time_slot": ts, "mid_tag": mid})
        enrollments.append((roll.upper(), short, section))

    # Wipe master data + request data for a clean reload
    for coll in ["students", "courses", "sections", "enrollments", "requests", "notifications", "import_staging"]:
        await db[coll].delete_many({})

    # Insert courses
    course_id_by_short = {}
    for short, c in courses.items():
        cid = f"course_{uuid.uuid4().hex[:10]}"
        course_id_by_short[short] = cid
        await db.courses.insert_one({"course_id": cid, **c})
    # Insert sections
    section_id_by_key = {}
    for (short, section), s in sections.items():
        sid = f"section_{uuid.uuid4().hex[:10]}"
        section_id_by_key[(short, section)] = sid
        await db.sections.insert_one({"section_id": sid, "course_id": course_id_by_short[short],
                                      "min_capacity": None, "max_capacity": None, **s})
    # Insert students
    for roll, st in students.items():
        await db.students.insert_one(st)
    # Insert enrollments
    ecount = 0
    for roll, short, section in enrollments:
        cid = course_id_by_short.get(short); sid = section_id_by_key.get((short, section))
        if not cid or not sid:
            continue
        await db.enrollments.insert_one({"pgpid": roll, "course_id": cid, "section_id": sid})
        ecount += 1

    await audit("IMPORT_TERMV", admin["email"], f"students={len(students)} courses={len(courses)} sections={len(sections)} enrollments={ecount}")
    return {"ok": True, "students": len(students), "courses": len(courses), "sections": len(sections), "enrollments": ecount}

# ------------------------------------------------------------------
# Trading Board (public marketplace of add/drop cases)
# ------------------------------------------------------------------
async def get_trading_settings():
    doc = await db.settings.find_one({"key": "trading"}, {"_id": 0})
    if not doc:
        doc = {"key": "trading", "enabled": True}
        await db.settings.insert_one(doc)
        doc.pop("_id", None)
    return doc

async def trading_is_open():
    t = await get_trading_settings()
    w = compute_window_state(await get_window_doc())
    return (t["enabled"] and w["is_open"]), t, w

def resolve_courses(ids, cmap):
    out = []
    for cid in ids:
        c = cmap.get(cid)
        if c:
            out.append({"course_id": cid, "course_code": c["course_code"], "course_name": c["course_name"], "credits": c.get("credits")})
    return out

def resolve_sections(ids, cmap, smap):
    out = []
    for sid in ids:
        s = smap.get(sid)
        if not s:
            continue
        c = cmap.get(s["course_id"], {})
        out.append({"section_id": sid, "course_id": s["course_id"], "course_code": c.get("course_code"),
                    "course_name": c.get("course_name"), "section_name": s.get("section_name"),
                    "day": s.get("day"), "time_slot": s.get("time_slot"), "credits": c.get("credits")})
    return out

class TradingPostInput(BaseModel):
    drop_course_ids: List[str] = []
    add_section_ids: List[str] = []
    note: Optional[str] = None

@api_router.get("/trading/board")
async def trading_board(student=Depends(require_student)):
    is_open, t, w = await trading_is_open()
    if not is_open:
        msg = "The trading board is currently closed by the administrator." if not t["enabled"] else w["message"]
        return {"enabled": False, "message": msg, "posts": []}
    cmap, smap = await build_course_section_maps()
    posts = await db.trading_posts.find({"active": True}, {"_id": 0}).sort("updated_at", -1).to_list(5000)
    out = [{
        "post_id": p["post_id"], "pgpid": p["pgpid"], "student_name": p["student_name"],
        "drop_courses": resolve_courses(p.get("drop_course_ids", []), cmap),
        "add_sections": resolve_sections(p.get("add_section_ids", []), cmap, smap),
        "note": p.get("note"), "updated_at": p["updated_at"], "is_mine": p["pgpid"] == student["pgpid"],
    } for p in posts]
    return {"enabled": True, "message": "", "posts": out}

@api_router.get("/trading/mine")
async def trading_mine(student=Depends(require_student)):
    p = await db.trading_posts.find_one({"pgpid": student["pgpid"], "active": True}, {"_id": 0})
    if not p:
        return {"post": None}
    return {"post": {"post_id": p["post_id"], "drop_course_ids": p.get("drop_course_ids", []),
                     "add_section_ids": p.get("add_section_ids", []), "note": p.get("note")}}

@api_router.post("/trading/posts")
async def upsert_trading_post(payload: TradingPostInput, student=Depends(require_student)):
    is_open, t, w = await trading_is_open()
    if not is_open:
        raise HTTPException(status_code=403, detail="The trading board is currently closed.")
    pgpid = student["pgpid"]
    enr = await db.enrollments.find({"pgpid": pgpid}, {"_id": 0}).to_list(1000)
    owned_courses = {e["course_id"] for e in enr}
    owned_sections = {e["section_id"] for e in enr}
    cmap, smap = await build_course_section_maps()
    for cid in payload.drop_course_ids:
        if cid not in owned_courses:
            raise HTTPException(status_code=400, detail="Under 'want to drop' you can only list courses you are currently enrolled in.")
    for sid in payload.add_section_ids:
        if sid not in smap or sid in owned_sections:
            raise HTTPException(status_code=400, detail="Under 'want to add' you can only list sections you don't already hold (a different section of a course you have is allowed).")
    if not payload.drop_course_ids and not payload.add_section_ids:
        raise HTTPException(status_code=400, detail="Add at least one course to drop or a section to add.")
    now = iso(now_utc())
    existing = await db.trading_posts.find_one({"pgpid": pgpid}, {"_id": 0})
    doc = {"pgpid": pgpid, "student_name": student["name"], "drop_course_ids": payload.drop_course_ids,
           "add_section_ids": payload.add_section_ids, "note": payload.note, "active": True, "updated_at": now}
    if existing:
        await db.trading_posts.update_one({"pgpid": pgpid}, {"$set": doc})
        post_id = existing["post_id"]
    else:
        post_id = f"trade_{uuid.uuid4().hex[:10]}"
        await db.trading_posts.insert_one({"post_id": post_id, "created_at": now, **doc})
    await audit("TRADING_POST", pgpid, "created/updated trading case")
    return {"ok": True, "post_id": post_id}

@api_router.delete("/trading/posts/{post_id}")
async def delete_trading_post(post_id: str, student=Depends(require_student)):
    p = await db.trading_posts.find_one({"post_id": post_id}, {"_id": 0})
    if not p or p["pgpid"] != student["pgpid"]:
        raise HTTPException(status_code=403, detail="Access denied")
    await db.trading_posts.delete_one({"post_id": post_id})
    await audit("TRADING_DELETE", student["pgpid"], "deleted own trading case", None)
    return {"ok": True}

@api_router.get("/admin/trading")
async def admin_trading(admin=Depends(require_admin)):
    t = await get_trading_settings()
    cmap, smap = await build_course_section_maps()
    posts = await db.trading_posts.find({}, {"_id": 0}).sort("updated_at", -1).to_list(5000)
    out = [{
        "post_id": p["post_id"], "pgpid": p["pgpid"], "student_name": p["student_name"],
        "drop_courses": resolve_courses(p.get("drop_course_ids", []), cmap),
        "add_sections": resolve_sections(p.get("add_section_ids", []), cmap, smap),
        "note": p.get("note"), "updated_at": p["updated_at"],
    } for p in posts]
    w = compute_window_state(await get_window_doc())
    return {"enabled": t["enabled"], "window_open": w["is_open"], "posts": out}

class TradingSettingsInput(BaseModel):
    enabled: bool

@api_router.put("/admin/trading/settings")
async def admin_trading_settings(payload: TradingSettingsInput, admin=Depends(require_admin)):
    await db.settings.update_one({"key": "trading"}, {"$set": {"enabled": payload.enabled}}, upsert=True)
    await audit("TRADING_SETTINGS", admin["email"], f"trading enabled={payload.enabled}")
    return {"ok": True, "enabled": payload.enabled}

@api_router.delete("/admin/trading")
async def admin_clear_trading(admin=Depends(require_admin)):
    res = await db.trading_posts.delete_many({})
    await audit("TRADING_ADMIN_CLEAR", admin["email"], f"cleared all trading posts ({res.deleted_count})")
    return {"ok": True, "deleted": res.deleted_count}

@api_router.delete("/admin/trading/{post_id}")
async def admin_delete_trading(post_id: str, admin=Depends(require_admin)):
    await db.trading_posts.delete_one({"post_id": post_id})
    await audit("TRADING_ADMIN_DELETE", admin["email"], f"removed trading post {post_id}")
    return {"ok": True}

@api_router.get("/")
async def root():
    return {"message": "PGP Course Change Request Portal API"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origin_regex=".*",
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
