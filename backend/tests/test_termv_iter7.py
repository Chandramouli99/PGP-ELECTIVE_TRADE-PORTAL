"""
Iteration 7: EXECUTION -> enrollment mutation tests.

Covers:
  - ADD executed -> enrollment appears in student dashboard + timetable
  - DROP executed -> enrollment removed
  - ADD_DROP executed -> drop removed, add present
  - approval alone (APPROVED_PENDING_EXECUTION) does NOT change enrollments
  - executed on non-approved => 400
  - SECTION_SWAP executed -> both students' sections swapped
  - COURSE_SWAP executed -> both students' course/section holdings exchanged
  - GET /admin/export/executed => valid xlsx with expected columns; student => 403
  - GET /admin/export still works
  - POST /admin/import/termv re-imports master data successfully (also our final cleanup)

State restoration: We snapshot PGP41071 and PGP41034 enrollments beforehand and at the
end of the module re-run POST /admin/import/termv to fully restore master data.
"""

import os
import io
import time
import requests
import pytest
import pymongo
from openpyxl import load_workbook

BASE = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
assert BASE, "REACT_APP_BACKEND_URL missing"
API = f"{BASE}/api"

STU = "stu_real"                    # PGP41071 (Chinmay Joshi)
STU_PGPID = "PGP41071"
ADMIN = "admintest_session"

PARTNER_PGPID = "PGP41034"
PARTNER_TOKEN = "stu_41034_iter7"

TERMV_XLSX = "/tmp/termv.xlsx"

MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "test_database")
mclient = pymongo.MongoClient(MONGO_URL)
mdb = mclient[DB_NAME]


def course_by_code(code):
    c = mdb.courses.find_one({"course_code": code})
    assert c, f"course {code} missing"
    return c["course_id"]


def section_by(course_id, section_name):
    s = mdb.sections.find_one({"course_id": course_id, "section_name": section_name})
    assert s, f"section {section_name} of {course_id} missing"
    return s["section_id"]


def sections_of(course_id):
    return list(mdb.sections.find({"course_id": course_id}, {"_id": 0}))


def h(tok):
    return {"Authorization": f"Bearer {tok}"}


def get_courses(tok):
    r = requests.get(f"{API}/student/dashboard", headers=h(tok), timeout=30)
    r.raise_for_status()
    return r.json()["courses"]


def has_course(tok, course_id, section_id=None):
    for c in get_courses(tok):
        if c["course_id"] == course_id and (section_id is None or c["section_id"] == section_id):
            return True
    return False


@pytest.fixture(scope="module", autouse=True)
def setup_partner_session_and_cleanup():
    # ---- setup partner user + session ----
    student = mdb.students.find_one({"pgpid": PARTNER_PGPID})
    assert student, f"partner {PARTNER_PGPID} not in master data"
    email = student["email"].lower()
    existing = mdb.users.find_one({"email": email})
    if existing:
        user_id = existing["user_id"]
        mdb.users.update_one({"user_id": user_id}, {"$set": {"role": "student", "pgpid": PARTNER_PGPID, "active": True}})
    else:
        user_id = f"user_iter7_{PARTNER_PGPID.lower()}"
        mdb.users.insert_one({
            "user_id": user_id, "email": email, "name": student["name"],
            "role": "student", "pgpid": PARTNER_PGPID, "active": True,
        })
    mdb.user_sessions.delete_many({"session_token": PARTNER_TOKEN})
    mdb.user_sessions.insert_one({
        "user_id": user_id, "session_token": PARTNER_TOKEN,
        "expires_at": "2099-01-01T00:00:00+00:00",
    })

    # snapshot enrollments for restoration
    snap_stu = list(mdb.enrollments.find({"pgpid": STU_PGPID}, {"_id": 0}))
    snap_par = list(mdb.enrollments.find({"pgpid": PARTNER_PGPID}, {"_id": 0}))
    yield
    # ---- teardown: purge test requests + re-import termv (full reset) ----
    mdb.requests.delete_many({"$or": [
        {"student_pgpid": STU_PGPID},
        {"student_pgpid": PARTNER_PGPID},
        {"swap.partner_pgpid": PARTNER_PGPID},
        {"swap.partner_pgpid": STU_PGPID},
    ]})
    mdb.trading_posts.delete_many({"pgpid": {"$in": [STU_PGPID, PARTNER_PGPID]}})
    mdb.notifications.delete_many({"pgpid": {"$in": [STU_PGPID, PARTNER_PGPID]}})
    mdb.user_sessions.delete_many({"session_token": PARTNER_TOKEN})

    # Best-effort direct restore in case import file missing
    mdb.enrollments.delete_many({"pgpid": {"$in": [STU_PGPID, PARTNER_PGPID]}})
    if snap_stu:
        mdb.enrollments.insert_many(snap_stu)
    if snap_par:
        mdb.enrollments.insert_many(snap_par)

    # Also re-run import as belt+braces if file exists
    if os.path.exists(TERMV_XLSX):
        try:
            with open(TERMV_XLSX, "rb") as f:
                requests.post(f"{API}/admin/import/termv",
                              headers=h(ADMIN),
                              files={"file": ("termv.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                              timeout=120)
        except Exception:
            pass


# ---------- helpers to drive request lifecycle ----------

def submit(payload, tok=STU):
    r = requests.post(f"{API}/student/requests", headers=h(tok), json=payload, timeout=30)
    return r


def admin_decide(req_id, decision, expect=200):
    r = requests.post(f"{API}/admin/requests/{req_id}/decision",
                      headers=h(ADMIN), json={"decision": decision}, timeout=30)
    assert r.status_code == expect, f"decision {decision} on {req_id}: {r.status_code} {r.text}"
    return r


def cleanup_requests():
    mdb.requests.delete_many({"$or": [
        {"student_pgpid": STU_PGPID}, {"student_pgpid": PARTNER_PGPID},
        {"swap.partner_pgpid": PARTNER_PGPID}, {"swap.partner_pgpid": STU_PGPID},
    ]})


# ============ Feature: EXECUTED export & admin export ============

class TestExports:
    def test_admin_export_all_requests(self):
        r = requests.get(f"{API}/admin/export", headers=h(ADMIN), timeout=30)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")

    def test_admin_export_executed_shape(self):
        r = requests.get(f"{API}/admin/export/executed", headers=h(ADMIN), timeout=30)
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        expected = ["PGPID", "Student Name", "Change", "Course", "From Section", "To Section", "Request ID", "Executed At"]
        assert headers_row == expected, f"headers were {headers_row}"

    def test_student_cannot_export_executed(self):
        r = requests.get(f"{API}/admin/export/executed", headers=h(STU), timeout=30)
        assert r.status_code == 403


# ============ Feature: ADD execution ============

class TestAddExecution:
    def setup_method(self):
        cleanup_requests()
        self.ADD_COURSE = course_by_code("BDA")
        self.ADD_SECTION = section_by(self.ADD_COURSE, "A")
        mdb.enrollments.delete_one({"pgpid": STU_PGPID, "course_id": self.ADD_COURSE})

    def teardown_method(self):
        mdb.enrollments.delete_one({"pgpid": STU_PGPID, "course_id": self.ADD_COURSE})
        cleanup_requests()

    def test_add_executed_appears_in_dashboard_and_timetable(self):
        r = submit({"request_type": "ADD", "add_course_id": self.ADD_COURSE, "add_section_id": self.ADD_SECTION})
        assert r.status_code == 200, r.text
        rid = r.json()["request_id"]

        # approve does NOT mutate enrollment
        assert not has_course(STU, self.ADD_COURSE)
        admin_decide(rid, "approve")
        assert not has_course(STU, self.ADD_COURSE), "approval must not modify enrollment"

        # executed mutates
        admin_decide(rid, "executed")
        assert has_course(STU, self.ADD_COURSE, self.ADD_SECTION), "ADD executed must show in dashboard"

        # timetable also shows new course
        tt = requests.get(f"{API}/student/timetable", headers=h(STU), timeout=30).json()
        # timetable is a schedule with slots referencing course_ids; do a substring search on json
        assert self.ADD_COURSE in str(tt), "ADD course must appear in timetable"

    def test_executed_on_non_approved_returns_400(self):
        r = submit({"request_type": "ADD", "add_course_id": self.ADD_COURSE, "add_section_id": self.ADD_SECTION})
        assert r.status_code == 200
        rid = r.json()["request_id"]
        r2 = requests.post(f"{API}/admin/requests/{rid}/decision",
                           headers=h(ADMIN), json={"decision": "executed"}, timeout=30)
        assert r2.status_code == 400


# ============ Feature: DROP execution ============

class TestDropExecution:
    def setup_method(self):
        cleanup_requests()
        self.DROP_COURSE = course_by_code("ACDM")
        self.DROP_SECTION = section_by(self.DROP_COURSE, "A")
        # ensure enrollment exists
        if not mdb.enrollments.find_one({"pgpid": STU_PGPID, "course_id": self.DROP_COURSE}):
            mdb.enrollments.insert_one({"pgpid": STU_PGPID, "course_id": self.DROP_COURSE, "section_id": self.DROP_SECTION})

    def teardown_method(self):
        # restore
        mdb.enrollments.delete_one({"pgpid": STU_PGPID, "course_id": self.DROP_COURSE})
        mdb.enrollments.insert_one({"pgpid": STU_PGPID, "course_id": self.DROP_COURSE, "section_id": self.DROP_SECTION})
        cleanup_requests()

    def test_drop_executed_removes_enrollment(self):
        r = submit({"request_type": "DROP", "drop_course_id": self.DROP_COURSE, "drop_section_id": self.DROP_SECTION})
        assert r.status_code == 200, r.text
        rid = r.json()["request_id"]
        admin_decide(rid, "approve")
        assert has_course(STU, self.DROP_COURSE), "approval alone must not drop enrollment"
        admin_decide(rid, "executed")
        assert not has_course(STU, self.DROP_COURSE), "DROP executed must remove enrollment"
        tt = requests.get(f"{API}/student/timetable", headers=h(STU), timeout=30).json()
        assert self.DROP_COURSE not in str(tt), "dropped course should be absent from timetable"


# ============ Feature: ADD_DROP execution ============

class TestAddDropExecution:
    def setup_method(self):
        cleanup_requests()
        self.DROP_COURSE = course_by_code("AOC")
        self.DROP_SECTION = section_by(self.DROP_COURSE, "A")
        self.ADD_COURSE = course_by_code("BDA")
        self.ADD_SECTION = section_by(self.ADD_COURSE, "A")
        # ensure drop exists, add does not
        if not mdb.enrollments.find_one({"pgpid": STU_PGPID, "course_id": self.DROP_COURSE}):
            mdb.enrollments.insert_one({"pgpid": STU_PGPID, "course_id": self.DROP_COURSE, "section_id": self.DROP_SECTION})
        mdb.enrollments.delete_one({"pgpid": STU_PGPID, "course_id": self.ADD_COURSE})

    def teardown_method(self):
        mdb.enrollments.delete_many({"pgpid": STU_PGPID, "course_id": {"$in": [self.ADD_COURSE, self.DROP_COURSE]}})
        mdb.enrollments.insert_one({"pgpid": STU_PGPID, "course_id": self.DROP_COURSE, "section_id": self.DROP_SECTION})
        cleanup_requests()

    def test_add_drop_executed(self):
        r = submit({"request_type": "ADD_DROP",
                    "drop_course_id": self.DROP_COURSE, "drop_section_id": self.DROP_SECTION,
                    "add_course_id": self.ADD_COURSE, "add_section_id": self.ADD_SECTION})
        assert r.status_code == 200, r.text
        rid = r.json()["request_id"]
        admin_decide(rid, "approve")
        admin_decide(rid, "executed")
        assert not has_course(STU, self.DROP_COURSE), "drop must be removed"
        assert has_course(STU, self.ADD_COURSE, self.ADD_SECTION), "add must be present"


# ============ Feature: SECTION_SWAP execution ============

class TestSectionSwapExecution:
    def setup_method(self):
        cleanup_requests()
        self.COURSE = course_by_code("Beh.Fin")
        self.STU_SEC = section_by(self.COURSE, "POST-A")
        self.PAR_SEC = section_by(self.COURSE, "POST-B")
        # ensure current enrollments
        mdb.enrollments.delete_many({"pgpid": {"$in": [STU_PGPID, PARTNER_PGPID]}, "course_id": self.COURSE})
        mdb.enrollments.insert_one({"pgpid": STU_PGPID, "course_id": self.COURSE, "section_id": self.STU_SEC})
        mdb.enrollments.insert_one({"pgpid": PARTNER_PGPID, "course_id": self.COURSE, "section_id": self.PAR_SEC})

    def teardown_method(self):
        mdb.enrollments.delete_many({"pgpid": {"$in": [STU_PGPID, PARTNER_PGPID]}, "course_id": self.COURSE})
        mdb.enrollments.insert_one({"pgpid": STU_PGPID, "course_id": self.COURSE, "section_id": self.STU_SEC})
        mdb.enrollments.insert_one({"pgpid": PARTNER_PGPID, "course_id": self.COURSE, "section_id": self.PAR_SEC})
        cleanup_requests()

    def test_section_swap_executed_swaps_both(self):
        r = submit({"request_type": "SECTION_SWAP",
                    "partner_pgpid": PARTNER_PGPID,
                    "swap_course_id": self.COURSE,
                    "my_section_id": self.STU_SEC,
                    "requested_section_id": self.PAR_SEC})
        assert r.status_code == 200, r.text
        rid = r.json()["request_id"]

        # partner confirms
        r2 = requests.post(f"{API}/student/swaps/{rid}/respond",
                           headers=h(PARTNER_TOKEN), json={"action": "accept"}, timeout=30)
        assert r2.status_code == 200, r2.text
        admin_decide(rid, "approve")
        # both sides unchanged still
        assert has_course(STU, self.COURSE, self.STU_SEC)
        assert has_course(PARTNER_TOKEN, self.COURSE, self.PAR_SEC)
        admin_decide(rid, "executed")
        # verify swapped
        assert has_course(STU, self.COURSE, self.PAR_SEC), "initiator moved to partner section"
        assert has_course(PARTNER_TOKEN, self.COURSE, self.STU_SEC), "partner moved to initiator section"


# ============ Feature: COURSE_SWAP execution ============

class TestCourseSwapExecution:
    def setup_method(self):
        cleanup_requests()
        self.GIVE_COURSE = course_by_code("ACDM")
        self.GIVE_SEC = section_by(self.GIVE_COURSE, "A")
        self.WANT_COURSE = course_by_code("BPIM-J")
        self.WANT_SEC = section_by(self.WANT_COURSE, "A")
        mdb.enrollments.delete_many({"pgpid": {"$in": [STU_PGPID, PARTNER_PGPID]}, "course_id": {"$in": [self.GIVE_COURSE, self.WANT_COURSE]}})
        mdb.enrollments.insert_one({"pgpid": STU_PGPID, "course_id": self.GIVE_COURSE, "section_id": self.GIVE_SEC})
        mdb.enrollments.insert_one({"pgpid": PARTNER_PGPID, "course_id": self.WANT_COURSE, "section_id": self.WANT_SEC})

    def teardown_method(self):
        mdb.enrollments.delete_many({"pgpid": {"$in": [STU_PGPID, PARTNER_PGPID]}, "course_id": {"$in": [self.GIVE_COURSE, self.WANT_COURSE]}})
        mdb.enrollments.insert_one({"pgpid": STU_PGPID, "course_id": self.GIVE_COURSE, "section_id": self.GIVE_SEC})
        mdb.enrollments.insert_one({"pgpid": PARTNER_PGPID, "course_id": self.WANT_COURSE, "section_id": self.WANT_SEC})
        cleanup_requests()

    def test_course_swap_executed_swaps_both(self):
        r = submit({"request_type": "COURSE_SWAP",
                    "partner_pgpid": PARTNER_PGPID,
                    "give_course_id": self.GIVE_COURSE, "give_section_id": self.GIVE_SEC,
                    "want_course_id": self.WANT_COURSE, "want_section_id": self.WANT_SEC})
        assert r.status_code == 200, r.text
        rid = r.json()["request_id"]
        r2 = requests.post(f"{API}/student/swaps/{rid}/respond",
                           headers=h(PARTNER_TOKEN), json={"action": "accept"}, timeout=30)
        assert r2.status_code == 200, r2.text
        admin_decide(rid, "approve")
        admin_decide(rid, "executed")
        # initiator: no longer holds GIVE, now holds WANT
        assert not has_course(STU, self.GIVE_COURSE), "initiator no longer holds offered course"
        assert has_course(STU, self.WANT_COURSE, self.WANT_SEC), "initiator now holds wanted course"
        # partner: no longer holds WANT, now holds GIVE
        assert not has_course(PARTNER_TOKEN, self.WANT_COURSE), "partner no longer holds requested course"
        assert has_course(PARTNER_TOKEN, self.GIVE_COURSE, self.GIVE_SEC), "partner now holds offered course"


# ============ Feature: Executed export contains rows after execution ============

class TestExecutedExportRows:
    def setup_method(self):
        cleanup_requests()
        self.ADD_COURSE = course_by_code("BDA")
        self.ADD_SECTION = section_by(self.ADD_COURSE, "A")
        mdb.enrollments.delete_one({"pgpid": STU_PGPID, "course_id": self.ADD_COURSE})

    def teardown_method(self):
        mdb.enrollments.delete_one({"pgpid": STU_PGPID, "course_id": self.ADD_COURSE})
        cleanup_requests()

    def test_executed_export_lists_our_execution(self):
        r = submit({"request_type": "ADD", "add_course_id": self.ADD_COURSE, "add_section_id": self.ADD_SECTION})
        rid = r.json()["request_id"]
        admin_decide(rid, "approve")
        admin_decide(rid, "executed")

        r = requests.get(f"{API}/admin/export/executed", headers=h(ADMIN), timeout=30)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
        assert any(rid == row[6] and row[2] == "ADD" for row in rows), f"expected ADD row for {rid} in export; got {rows[-5:]}"


# ============ Feature: Master re-upload ============

class TestMasterReupload:
    def test_import_termv_succeeds(self):
        if not os.path.exists(TERMV_XLSX):
            pytest.skip("no termv.xlsx available")
        with open(TERMV_XLSX, "rb") as f:
            r = requests.post(f"{API}/admin/import/termv",
                              headers=h(ADMIN),
                              files={"file": ("termv.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                              timeout=180)
        assert r.status_code == 200, r.text
        data = r.json()
        # Basic sanity: counts present
        assert any(k in data for k in ("students", "courses", "sections", "enrollments", "counts", "message", "ok"))
